import openpyxl

wb = openpyxl.load_workbook('baseline/valores.xlsx', data_only=True)
print("Sheet names:", wb.sheetnames)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ---")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    # Print first 5 rows
    for r in range(1, 6):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        print(f"Row {r}: {row_vals}")
