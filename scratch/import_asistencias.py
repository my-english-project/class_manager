import openpyxl
import mysql.connector

# Read env variables
env = {}
with open('.env', 'r') as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        k, v = line.split('=', 1)
        env[k.strip()] = v.strip()

conn = mysql.connector.connect(
    host=env.get('DB_HOST', 'localhost'),
    port=int(env.get('DB_PORT', 3306)),
    user=env.get('DB_USER', 'root'),
    password=env.get('DB_PASS', 'kjb2a0p'),
    database=env.get('DB_NAME', 'uts')
)
cursor = conn.cursor(dictionary=True)

# 1. Map columns to dates and partials
months_map = {
    'Enero': 1,
    'Febrero': 2,
    'Marzo': 3,
    'Abril': 4
}

wb = openpyxl.load_workbook('baseline/asistencias.xlsx', data_only=True)
ws = wb.active

current_month = None
col_info = {}

col_counter = 1
for c in range(3, ws.max_column + 1):
    m_val = ws.cell(row=1, column=c).value
    if m_val is not None:
        current_month = months_map[m_val.strip()]
    d_val = ws.cell(row=2, column=c).value
    date_str = f"2026-{current_month:02d}-{int(d_val):02d}"
    
    if col_counter <= 13:
        parcial = 1
    elif col_counter <= 26:
        parcial = 2
    else:
        parcial = 3
        
    col_info[c] = {
        'date': date_str,
        'parcial': parcial
    }
    col_counter += 1

# 2. Get students and their groups from DB
cursor.execute("""
    SELECT a.id_alumno, a.matricula, ga.id_grupo
    FROM alumno a
    INNER JOIN grupo_alumno ga ON a.id_alumno = ga.id_alumno
""")
db_alumnos = cursor.fetchall()
db_map = {int(a['matricula']): a for a in db_alumnos}

# 3. Create sessions for groups in the database
cursor.execute("SELECT id_grupo FROM grupo WHERE ciclo = '26C1'")
groups_26c1 = [g['id_grupo'] for g in cursor.fetchall()]

session_map = {}
for group_id in groups_26c1:
    for c, info in col_info.items():
        date_str = info['date']
        parcial = info['parcial']
        
        cursor.execute("SELECT id_sesion FROM sesion WHERE id_grupo = %s AND fecha = %s", (group_id, date_str))
        res = cursor.fetchone()
        if res:
            session_id = res['id_sesion']
        else:
            cursor.execute("""
                INSERT INTO sesion (id_grupo, fecha, parcial, tema)
                VALUES (%s, %s, %s, %s)
            """, (group_id, date_str, parcial, "Clase importada"))
            session_id = cursor.lastrowid
            
        session_map[(group_id, date_str)] = session_id

# 4. Insert attendances
total_records = 0
for r in range(3, ws.max_row + 1):
    enrollment_val = ws.cell(row=r, column=1).value
    if enrollment_val is None:
        continue
    enrollment = int(enrollment_val)
    
    if enrollment not in db_map:
        print(f"Warning: Enrollment {enrollment} not found in DB!")
        continue
        
    student_db = db_map[enrollment]
    student_id = student_db['id_alumno']
    group_id = student_db['id_grupo']
    
    for c in range(3, ws.max_column + 1):
        date_str = col_info[c]['date']
        session_id = session_map.get((group_id, date_str))
        
        if not session_id:
            continue
            
        cell_val = ws.cell(row=r, column=c).value
        estado = 'asistencia' if cell_val else 'falta'
        
        cursor.execute("""
            INSERT INTO asistencia (id_sesion, id_alumno, estado)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE estado = VALUES(estado)
        """, (session_id, student_id, estado))
        total_records += 1

conn.commit()
print(f"Successfully processed and stored {total_records} attendance records across sessions for groups in cycle 26C1!")

cursor.close()
conn.close()
