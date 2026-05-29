import openpyxl
import mysql.connector
import os

# Read env variables
env = {}
with open('.env', 'r') as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        k, v = line.split('=', 1)
        env[k.strip()] = v.strip()

conn = mysql.connector.connect(
    host=env.get('DB_HOST', 'localhost'),
    port=int(env.get('DB_PORT', 3306)),
    user=env.get('DB_USER', 'root'),
    password=env.get('DB_PASS', 'kjb2a0p'),
    database=env.get('DB_NAME', 'uts')
)
cursor = conn.cursor(dictionary=True)

# Get all students and their groups
cursor.execute("""
    SELECT a.id_alumno, a.matricula, ga.id_grupo
    FROM alumno a
    INNER JOIN grupo_alumno ga ON a.id_alumno = ga.id_alumno
""")
db_alumnos = cursor.fetchall()

db_map = {int(a['matricula']): a for a in db_alumnos}

wb = openpyxl.load_workbook('baseline/asistencias.xlsx', data_only=True)
ws = wb.active

print(f"Total sheet rows: {ws.max_row}")
matched = 0
unmatched = []

for r in range(3, ws.max_row + 1):
    enrollment = ws.cell(row=r, column=1).value
    name = ws.cell(row=r, column=2).value
    if enrollment is None:
        continue
    
    enrollment = int(enrollment)
    if enrollment in db_map:
        db_student = db_map[enrollment]
        matched += 1
        print(f"Row {r}: Matched {enrollment} ({name}) to Group {db_student['id_grupo']}, Student ID {db_student['id_alumno']}")
    else:
        unmatched.append((r, enrollment, name))

print(f"\nSummary: Matched {matched} students.")
if unmatched:
    print("Unmatched students:")
    for r, e, n in unmatched:
        print(f"Row {r}: {e} - {n}")
else:
    print("All students matched perfectly!")

cursor.close()
conn.close()
