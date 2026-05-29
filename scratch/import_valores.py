import openpyxl
import mysql.connector

# Read env variables
env = {}
with open('.env', 'r') as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        k, v = line.split('=', 1)
        env[k.strip()] = v.strip()

conn = mysql.connector.connect(
    host=env.get('DB_HOST', 'localhost'),
    port=int(env.get('DB_PORT', 3306)),
    user=env.get('DB_USER', 'root'),
    password=env.get('DB_PASS', 'kjb2a0p'),
    database=env.get('DB_NAME', 'uts')
)
cursor = conn.cursor(dictionary=True)

# 1. Get students and their groups from DB
cursor.execute("""
    SELECT a.id_alumno, a.matricula, ga.id_grupo
    FROM alumno a
    INNER JOIN grupo_alumno ga ON a.id_alumno = ga.id_alumno
""")
db_alumnos = cursor.fetchall()
db_map = {int(a['matricula']): a for a in db_alumnos}

# List of first 5 students of IER8A (Group 1) to skip in Write exam
skip_matriculas = {612310503, 612310220, 612310550, 612310216, 612310272}

# Clear any pre-existing Write exam grades for the 5 skipped students
cursor.execute("""
    DELETE FROM examen_escrito
    WHERE id_grupo = 1 AND id_alumno IN (
        SELECT id_alumno FROM alumno WHERE matricula IN (612310503, 612310220, 612310550, 612310216, 612310272)
    )
""")
conn.commit()

wb = openpyxl.load_workbook('baseline/valores.xlsx', data_only=True)

# Helper function to get or create unique activity for portfolio / homeworks
activity_cache = {}
def get_or_create_activity(group_id, tipo, parcial):
    key = (group_id, tipo, parcial)
    if key in activity_cache:
        return activity_cache[key]
        
    db_name = "Portafolio Único" if tipo == 'portafolio' else "Tarea Única"
    cursor.execute("""
        SELECT id_actividad FROM actividad
        WHERE id_grupo = %s AND tipo = %s AND parcial = %s
    """, (group_id, tipo, parcial))
    res = cursor.fetchone()
    if res:
        act_id = res['id_actividad']
    else:
        cursor.execute("""
            INSERT INTO actividad (id_grupo, tipo, parcial, nombre, orden)
            VALUES (%s, %s, %s, %s, 1)
        """, (group_id, tipo, parcial, db_name))
        act_id = cursor.lastrowid
        
    activity_cache[key] = act_id
    return act_id

# Track statistics
counts = {
    'Write exam': 0,
    'Oral exam': 0,
    'Portfolio': 0,
    'Homeworks': 0,
    'skipped_write': 0
}

# --- PROCESS WRITE EXAM ---
ws_write = wb['Write exam']
for r in range(2, ws_write.max_row + 1):
    matricula_val = ws_write.cell(row=r, column=1).value
    if matricula_val is None:
        continue
    matricula = int(matricula_val)
    
    if matricula not in db_map:
        continue
    student = db_map[matricula]
    
    # Check if student is IER8A and in first 5 skip list
    if student['id_grupo'] == 1 and matricula in skip_matriculas:
        counts['skipped_write'] += 3 # skip all 3 partials
        continue
        
    for p in [1, 2, 3]:
        col_idx = p + 2
        score = ws_write.cell(row=r, column=col_idx).value
        if score is not None:
            cursor.execute("""
                INSERT INTO examen_escrito (id_grupo, id_alumno, parcial, calificacion)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE calificacion = VALUES(calificacion)
            """, (student['id_grupo'], student['id_alumno'], p, float(score)))
            counts['Write exam'] += 1

# --- PROCESS ORAL EXAM ---
ws_oral = wb['Oral exam']
for r in range(2, ws_oral.max_row + 1):
    matricula_val = ws_oral.cell(row=r, column=1).value
    if matricula_val is None:
        continue
    matricula = int(matricula_val)
    
    if matricula not in db_map:
        continue
    student = db_map[matricula]
    
    for p in [1, 2, 3]:
        col_idx = p + 2
        score = ws_oral.cell(row=r, column=col_idx).value
        if score is not None:
            cursor.execute("""
                INSERT INTO examen_oral (id_grupo, id_alumno, parcial, calificacion)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE calificacion = VALUES(calificacion)
            """, (student['id_grupo'], student['id_alumno'], p, float(score)))
            counts['Oral exam'] += 1

# --- PROCESS PORTFOLIO ---
ws_port = wb['Portfolio']
for r in range(2, ws_port.max_row + 1):
    matricula_val = ws_port.cell(row=r, column=1).value
    if matricula_val is None:
        continue
    matricula = int(matricula_val)
    
    if matricula not in db_map:
        continue
    student = db_map[matricula]
    
    for p in [1, 2, 3]:
        col_idx = p + 2
        score = ws_port.cell(row=r, column=col_idx).value
        if score is not None:
            act_id = get_or_create_activity(student['id_grupo'], 'portafolio', p)
            cursor.execute("""
                INSERT INTO calificacion_actividad (id_actividad, id_alumno, calificacion)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE calificacion = VALUES(calificacion)
            """, (act_id, student['id_alumno'], float(score)))
            counts['Portfolio'] += 1

# --- PROCESS HOMEWORKS ---
ws_home = wb['Homeworks']
for r in range(2, ws_home.max_row + 1):
    matricula_val = ws_home.cell(row=r, column=1).value
    if matricula_val is None:
        continue
    matricula = int(matricula_val)
    
    if matricula not in db_map:
        continue
    student = db_map[matricula]
    
    for p in [1, 2, 3]:
        col_idx = p + 2
        score = ws_home.cell(row=r, column=col_idx).value
        if score is not None:
            act_id = get_or_create_activity(student['id_grupo'], 'tarea', p)
            cursor.execute("""
                INSERT INTO calificacion_actividad (id_actividad, id_alumno, calificacion)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE calificacion = VALUES(calificacion)
            """, (act_id, student['id_alumno'], float(score)))
            counts['Homeworks'] += 1

conn.commit()
print("Successfully imported academic grades and cleaned up skipped students!")
print(f"Write exam: {counts['Write exam']} records imported ({counts['skipped_write']} skipped)")
print(f"Oral exam: {counts['Oral exam']} records imported")
print(f"Portfolio: {counts['Portfolio']} records imported")
print(f"Homeworks: {counts['Homeworks']} records imported")

cursor.close()
conn.close()
