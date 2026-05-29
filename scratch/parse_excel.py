import openpyxl
import json
import re

wb = openpyxl.load_workbook('baseline/alumnos.xlsx')
sheet = wb.active

alumnos = []

def parse_mexican_name(full_name):
    # Clean non-breaking spaces
    full_name = full_name.replace('\xa0', ' ').strip()
    # Normalize spaces
    full_name = re.sub(r'\s+', ' ', full_name)
    words = full_name.split()
    if len(words) >= 3:
        paterno = words[0]
        materno = words[1]
        nombres = " ".join(words[2:])
    elif len(words) == 2:
        paterno = words[0]
        materno = ""
        nombres = words[1]
    else:
        paterno = words[0] if words else ""
        materno = ""
        nombres = ""
    return nombres, paterno, materno

for r in range(2, sheet.max_row + 1):
    mat = sheet.cell(r, 1).value
    name = sheet.cell(r, 2).value
    carrera = sheet.cell(r, 3).value
    grupo = sheet.cell(r, 4).value
    cuat = sheet.cell(r, 5).value
    ciclo = sheet.cell(r, 6).value
    materia = sheet.cell(r, 7).value
    
    if mat and name:
        nombres, paterno, materno = parse_mexican_name(str(name))
        alumnos.append({
            'matricula': str(mat).strip(),
            'nombre': nombres,
            'apellido_pat': paterno,
            'apellido_mat': materno,
            'carrera': str(carrera).strip() if carrera else "",
            'grupo': str(grupo).strip() if grupo else "A",
            'cuatrimestre': int(cuat) if cuat else 8,
            'ciclo': str(ciclo).strip() if ciclo else "26C1",
            'materia': str(materia).strip() if materia else "Inglés VIII"
        })

with open('scratch/alumnos.json', 'w', encoding='utf-8') as f:
    json.dump(alumnos, f, ensure_ascii=False, indent=2)

print(f"✓ parsed {len(alumnos)} students to scratch/alumnos.json")
